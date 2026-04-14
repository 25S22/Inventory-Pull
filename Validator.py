"""
ETL Script: Hostname Lookup from Resource Inventory
=====================================================
Accepts two Excel files:
  1. Resource Inventory  — master list of all hosts with details
  2. Verification Sheet  — hostnames to verify (one or more sheets)

Matches hostnames, extracts desired columns, and produces:
  • A formatted Excel report  (Found rows + Not Found rows + Summary)
  • An HTML email draft       (open in browser, copy into your mail client)

HOW TO RUN (no arguments needed if FILE paths are set in CONFIG below):
    python etl_hostname_lookup.py

CLI overrides (all optional — they take precedence over CONFIG values):
    python etl_hostname_lookup.py \
        --inventory   "Resource_Inventory.xlsx" \
        --verification "Verification.xlsx"      \
        --output      "Lookup_Results.xlsx"     \
        --email-draft "email_draft.html"

Edit only the CONFIG block below. Do not touch the code beneath it.
"""

# =============================================================================
#  CONFIGURATION  <-- All user-editable settings live here
# =============================================================================
CONFIG = {
    # File paths
    "inventory_file": "Resource_Inventory.xlsx",
    "verification_file": "Verification.xlsx",
    "output_file": "Lookup_Results.xlsx",
    "email_draft_file": "",

    # Inventory columns
    "inventory_hostname_col": "Instance Name",
    "inventory_desired_cols": [
        "Instance Name",
        "IP Address",
        "Environment",
        "OS",
        "Owner",
        "Status",
    ],
    "inventory_sheet": 0,

    # Verification columns
    "verification_hostname_col": "Name",
    "verification_all_sheets": True,
    "verification_sheet_names": ["Sheet1", "Sheet2"],

    # Matching behaviour
    "strip_before_at": True,
    "case_insensitive": True,

    # Output labels
    "not_found_message": "Hostname not found in the recent inventory",
    "status_col_name": "Lookup Status",
    "source_sheet_col_name": "Source Sheet",
    "original_value_col_name": "Original Verification Value",

    # Email draft
    "email_to": "recipient@example.com",
    "email_cc": "",
    "email_subject": "Hostname Verification Report",
    "email_body": (
        "Hi,\n\n"
        "Please find attached the hostname verification report generated from "
        "the latest resource inventory.\n\n"
        "The report contains:\n"
        "  * Found entries  - full inventory details for matched hostnames.\n"
        "  * Not Found entries - hostnames that could not be located in the "
        "current inventory.\n\n"
        "Kindly review and revert with any corrections.\n\n"
        "Regards"
    ),
}
# =============================================================================
#  END OF CONFIGURATION -- do not edit below this line
# =============================================================================

import argparse
import os
import sys
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# Style constants
_HEADER_FILL = PatternFill("solid", fgColor="1F3864")
_FOUND_FILL = PatternFill("solid", fgColor="E8F5E9")
_NOTFOUND_FILL = PatternFill("solid", fgColor="FFEBEE")
_ALT_ROW_FILL = PatternFill("solid", fgColor="F5F5F5")

_HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
_CELL_FONT = Font(name="Arial", size=10)
_BOLD_FONT = Font(name="Arial", bold=True, size=10)

_THIN = Side(style="thin", color="BDBDBD")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


# =============================================================================
#  Helpers
# =============================================================================

def _parse_hostname(raw: str, strip_at: bool) -> str:
    """Return the lookup key extracted from a raw cell value."""
    raw = str(raw).strip()
    if strip_at and "@" in raw:
        raw = raw.split("@", 1)[1].strip()
    return raw


def _normalise(value: str, case_insensitive: bool) -> str:
    s = str(value).strip()
    return s.lower() if case_insensitive else s


def _email_draft_path(output_path: str, explicit: str) -> str:
    """Derive the email draft path from the output path, or use the explicit one."""
    if explicit:
        return explicit
    p = Path(output_path)
    return str(p.with_name(p.stem + "_email_draft.html"))


def _read_tabular(path: str, sheet_name=0):
    """
    Read either Excel or CSV safely.

    Supported:
      - Excel: .xlsx .xls .xlsm .xlsb .ods
      - CSV:   .csv

    Behavior:
      - If sheet_name is None:
          * Excel -> dict[sheet_name, DataFrame]
          * CSV   -> {"Sheet1": DataFrame}
      - Else:
          * returns single DataFrame
    """
    p = Path(path)
    ext = p.suffix.lower()

    if ext == ".csv":
        df = pd.read_csv(path, dtype=str).fillna("")
        if sheet_name is None:
            return {"Sheet1": df}
        return df

    excel_exts = {".xlsx", ".xls", ".xlsm", ".xlsb", ".ods"}
    if ext in excel_exts:
        return pd.read_excel(path, sheet_name=sheet_name, dtype=str).fillna("")

    # Fallback by content/reader attempt
    try:
        df = pd.read_csv(path, dtype=str).fillna("")
        if sheet_name is None:
            return {"Sheet1": df}
        return df
    except Exception:
        return pd.read_excel(path, sheet_name=sheet_name, dtype=str).fillna("")


# =============================================================================
#  ETL steps
# =============================================================================

def load_inventory(path: str, cfg: dict) -> pd.DataFrame:
    """Load the inventory sheet/file and validate required columns."""
    inv = _read_tabular(path, sheet_name=cfg["inventory_sheet"])

    # Deduplicate: hostname col must not be checked twice if it also appears
    # in inventory_desired_cols (avoids confusing double-error messages).
    required = list(dict.fromkeys(
        [cfg["inventory_hostname_col"]] + cfg["inventory_desired_cols"]
    ))
    missing = [c for c in required if c not in inv.columns]
    if missing:
        print(
            "\n[ERROR] Columns not found in the inventory:"
            f"\n        {missing}"
            f"\n        Available columns: {list(inv.columns)}\n"
        )
        sys.exit(1)

    return inv


def build_lookup(inv: pd.DataFrame, cfg: dict) -> dict:
    """Build a dict: normalised hostname -> list[pd.Series] of matching rows."""
    lookup: dict = {}
    key_col = cfg["inventory_hostname_col"]
    ci = cfg["case_insensitive"]
    for _, row in inv.iterrows():
        key = _normalise(row[key_col], ci)
        if key:  # skip blank inventory hostnames
            lookup.setdefault(key, []).append(row)
    return lookup


def process_verification(vpath: str, cfg: dict, lookup: dict) -> list:
    """
    Walk every requested sheet/file in the verification input.
    Returns a flat list of result dicts, one dict per output row.
    """
    if cfg["verification_all_sheets"]:
        raw_sheets = _read_tabular(vpath, sheet_name=None)
        sheets = list(raw_sheets.items())
    else:
        sheets = [
            (name, _read_tabular(vpath, sheet_name=name))
            for name in cfg["verification_sheet_names"]
        ]

    hn_col = cfg["verification_hostname_col"]
    ci = cfg["case_insensitive"]
    desired = cfg["inventory_desired_cols"]
    not_found_msg = cfg["not_found_message"]
    src_col = cfg["source_sheet_col_name"]
    orig_col = cfg["original_value_col_name"]
    status_col = cfg["status_col_name"]

    results = []

    for sheet_name, df in sheets:
        df = df.fillna("")
        if hn_col not in df.columns:
            print(
                f"[WARNING] Sheet '{sheet_name}' has no column '{hn_col}'. "
                f"Available: {list(df.columns)}. Skipping."
            )
            continue

        for _, vrow in df.iterrows():
            raw_val = str(vrow[hn_col]).strip()
            if not raw_val:
                continue  # skip blank cells

            parsed = _parse_hostname(raw_val, cfg["strip_before_at"])
            key = _normalise(parsed, ci)

            if key in lookup:
                for inv_row in lookup[key]:
                    record = {src_col: sheet_name, orig_col: raw_val}
                    for col in desired:
                        record[col] = inv_row.get(col, "")
                    record[status_col] = "Found"
                    results.append(record)
            else:
                record = {src_col: sheet_name, orig_col: raw_val}
                for col in desired:
                    record[col] = ""
                record[status_col] = not_found_msg
                results.append(record)

    return results


# =============================================================================
#  Excel output
# =============================================================================

def _style_sheet(ws, df: pd.DataFrame, found_flags: list, cfg: dict) -> None:
    """Write df into ws with full formatting."""
    columns = list(df.columns)
    status_col_idx = columns.index(cfg["status_col_name"]) + 1  # 1-based

    # Header row
    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.border = _BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Data rows — zip positionally; no .loc / index dependency
    for row_pos, (row_values, is_found) in enumerate(
        zip(df.itertuples(index=False, name=None), found_flags), start=2
    ):
        base_fill = (
            (_ALT_ROW_FILL if row_pos % 2 == 0 else _FOUND_FILL)
            if is_found else _NOTFOUND_FILL
        )

        for col_idx, value in enumerate(row_values, start=1):
            cell = ws.cell(row=row_pos, column=col_idx, value=value)
            cell.font = _BOLD_FONT if col_idx == status_col_idx else _CELL_FONT
            cell.border = _BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=False)
            cell.fill = (
                (_FOUND_FILL if is_found else _NOTFOUND_FILL)
                if col_idx == status_col_idx
                else base_fill
            )

    # Auto column widths
    for col_idx, col_name in enumerate(columns, start=1):
        col_letter = get_column_letter(col_idx)
        cell_lengths = [
            len(str(ws.cell(r, col_idx).value or ""))
            for r in range(2, ws.max_row + 1)
        ]
        max_len = max([len(str(col_name))] + cell_lengths)
        ws.column_dimensions[col_letter].width = min(max_len + 4, 52)

    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"


def _write_summary(ws, df: pd.DataFrame, cfg: dict) -> None:
    status_col = cfg["status_col_name"]
    total = len(df)
    found_cnt = int((df[status_col] == "Found").sum())
    nf_cnt = total - found_cnt
    pct_found = f"{found_cnt / total * 100:.1f}%" if total else "N/A"

    ws["A1"] = "Hostname Verification -- Summary"
    ws["A1"].font = Font(name="Arial", bold=True, size=14, color="1F3864")
    ws["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d  %H:%M')}"
    ws["A2"].font = Font(name="Arial", size=10, color="757575")

    summary_rows = [
        ("Metric", "Value"),
        ("Total Hostnames Processed", total),
        ("Found in Inventory", found_cnt),
        ("Not Found in Inventory", nf_cnt),
        ("Match Rate", pct_found),
        ("", ""),
        ("Breakdown by Verification Sheet", ""),
    ]
    for sheet_name in df[cfg["source_sheet_col_name"]].unique():
        sub = df[df[cfg["source_sheet_col_name"]] == sheet_name]
        sf = int((sub[status_col] == "Found").sum())
        summary_rows.append((f"  {sheet_name}", f"{sf} / {len(sub)} found"))

    for r_idx, (label, value) in enumerate(summary_rows, start=4):
        c_label = ws.cell(row=r_idx, column=1, value=label)
        c_value = ws.cell(row=r_idx, column=2, value=value)
        if label == "Metric":
            for c in (c_label, c_value):
                c.font = _HEADER_FONT
                c.fill = _HEADER_FILL
                c.border = _BORDER
                c.alignment = Alignment(horizontal="center", vertical="center")
        elif label:
            c_label.font = _BOLD_FONT
            c_value.font = _CELL_FONT
            for c in (c_label, c_value):
                c.border = _BORDER

    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 22


def write_output_excel(results: list, output_path: str, cfg: dict) -> None:
    if not results:
        print("[WARNING] No results to write -- output file not created.")
        return

    df = pd.DataFrame(results)
    found_flags = [r == "Found" for r in df[cfg["status_col_name"]]]

    wb = Workbook()
    wb.remove(wb.active)

    ws_all = wb.create_sheet("All Results")
    _style_sheet(ws_all, df.reset_index(drop=True), found_flags, cfg)

    df_found = df[df[cfg["status_col_name"]] == "Found"].reset_index(drop=True)
    if not df_found.empty:
        ws_found = wb.create_sheet("Found")
        _style_sheet(ws_found, df_found, [True] * len(df_found), cfg)

    df_nf = df[df[cfg["status_col_name"]] != "Found"].reset_index(drop=True)
    if not df_nf.empty:
        ws_nf = wb.create_sheet("Not Found")
        _style_sheet(ws_nf, df_nf, [False] * len(df_nf), cfg)

    ws_sum = wb.create_sheet("Summary")
    _write_summary(ws_sum, df, cfg)

    wb.save(output_path)
    print(f"[OK] Report saved         -> {output_path}")


# =============================================================================
#  Email draft
# =============================================================================

def generate_email_draft(output_excel: str, results: list, cfg: dict, draft_path: str) -> None:
    status_col = cfg["status_col_name"]
    found = sum(1 for r in results if r[status_col] == "Found")
    nf = len(results) - found
    body = cfg["email_body"].replace("\n", "<br>")
    fname = os.path.basename(output_excel)
    cc_block = (
        f"<div class='field'><div class='label'>CC</div>"
        f"<div class='value'>{cfg['email_cc']}</div></div>"
        if cfg["email_cc"] else ""
    )

    html = f"""<!DOCTYPE html>
<html><head><meta charset="utf-8"><style>
body      {{font-family:Arial,sans-serif;margin:30px;color:#212121}}
.card     {{border:1px solid #e0e0e0;border-radius:10px;padding:26px;max-width:680px;background:#fff}}
.field    {{margin-bottom:12px}}
.label    {{font-size:11px;font-weight:bold;color:#757575;text-transform:uppercase;letter-spacing:.5px}}
.value    {{font-size:14px;color:#1a1a1a;padding:6px 0;border-bottom:1px solid #f0f0f0}}
.body-box {{background:#f9f9f9;border-left:4px solid #1F3864;padding:14px 18px;border-radius:4px;
            margin:14px 0;font-size:14px;line-height:1.75}}
.badge    {{display:inline-block;padding:4px 12px;border-radius:12px;font-size:12px;font-weight:bold;margin:2px}}
.g        {{background:#e8f5e9;color:#2e7d32}}
.r        {{background:#ffebee;color:#c62828}}
.b        {{background:#e8eaf6;color:#283593}}
.attach   {{font-size:13px;background:#e3f2fd;border:1px dashed #90caf9;padding:8px 14px;
            border-radius:6px;display:inline-block;margin-top:4px}}
h2        {{color:#1F3864;margin-bottom:4px;font-size:18px}}
.note     {{font-size:11px;color:#9e9e9e;margin-top:20px}}
</style></head><body>
<div class="card">
  <h2>&#128231; Email Draft</h2>
  <p style="font-size:12px;color:#9e9e9e;margin:0 0 16px">Open this file in a browser and copy the fields into your email client.</p>
  <div class="field"><div class="label">To</div><div class="value">{cfg["email_to"]}</div></div>
  {cc_block}
  <div class="field"><div class="label">Subject</div><div class="value">{cfg["email_subject"]}</div></div>
  <div class="field"><div class="label">Body</div><div class="body-box">{body}</div></div>
  <div class="field"><div class="label">Attachment</div><br><span class="attach">&#128206; {fname}</span></div>
  <div style="margin-top:20px">
    <span class="badge g">&#10004; Found: {found}</span>
    <span class="badge r">&#10008; Not Found: {nf}</span>
    <span class="badge b">Total: {found + nf}</span>
  </div>
  <p class="note">Generated by ETL Hostname Lookup &middot; {datetime.now().strftime("%Y-%m-%d %H:%M")}</p>
</div></body></html>"""

    with open(draft_path, "w", encoding="utf-8") as fh:
        fh.write(html)
    print(f"[OK] Email draft saved    -> {draft_path}")


# =============================================================================
#  CLI
# =============================================================================

def parse_args(cfg: dict) -> argparse.Namespace:
    p = argparse.ArgumentParser(
        description="ETL: Look up verification hostnames in a resource inventory.",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )
    p.add_argument(
        "--inventory",
        default=cfg["inventory_file"],
        help="Path to the Resource Inventory file (.xlsx/.xls/.csv).",
    )
    p.add_argument(
        "--verification",
        default=cfg["verification_file"],
        help="Path to the Verification file (.xlsx/.xls/.csv).",
    )
    p.add_argument(
        "--output",
        default=cfg["output_file"],
        help="Output Excel report path.",
    )
    p.add_argument(
        "--email-draft",
        default=cfg["email_draft_file"],
        help="HTML email draft path (auto-named next to --output if omitted).",
    )
    return p.parse_args()


def main() -> None:
    cfg = CONFIG
    args = parse_args(cfg)

    inventory_path = args.inventory
    verification_path = args.verification
    output_path = args.output
    draft_path = _email_draft_path(output_path, args.email_draft)

    # Validate input files
    errors = []
    for fpath, label in [
        (inventory_path, "Inventory"),
        (verification_path, "Verification"),
    ]:
        if not Path(fpath).is_file():
            errors.append(f"  [{label}] File not found: {fpath}")
    if errors:
        print("\n[ERROR] Cannot start -- input file(s) missing:")
        print("\n".join(errors))
        print("\nSet the correct paths in CONFIG or pass --inventory / --verification.")
        sys.exit(1)

    print()
    print("=" * 46)
    print("  ETL Hostname Lookup")
    print("=" * 46)
    print(f"  Inventory    : {inventory_path}")
    print(f"  Verification : {verification_path}")
    print(f"  Output       : {output_path}")
    print(f"  Email draft  : {draft_path}")
    print("=" * 46)
    print()

    print("[1/4] Loading resource inventory ...")
    inv = load_inventory(inventory_path, cfg)
    print(f"      {len(inv):,} rows loaded.")

    print("[2/4] Building lookup index ...")
    lookup = build_lookup(inv, cfg)
    print(f"      {len(lookup):,} unique hostnames indexed.")

    print("[3/4] Processing verification sheet(s) ...")
    results = process_verification(verification_path, cfg, lookup)
    found = sum(1 for r in results if r[cfg["status_col_name"]] == "Found")
    nf = len(results) - found
    print(f"      {len(results):,} rows -> {found:,} found, {nf:,} not found.")

    print("[4/4] Writing outputs ...")
    write_output_excel(results, output_path, cfg)
    generate_email_draft(output_path, results, cfg, draft_path)

    print()
    print("Done!")
    print()


if __name__ == "__main__":
    main()
