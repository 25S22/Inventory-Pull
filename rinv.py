"""
Apptio Cloudability – EC2 Resource Inventory Automation
========================================================
Pulls EC2 inventory from the Cloudability Resource Inventory Public API,
exports to a formatted .xlsx file, and dispatches it via Microsoft Outlook.

Architecture
------------
  POST  /v3/resource-inventory/jobs   → queue export job
  GET   /v3/resource-inventory/jobs/{id} → poll until FINISHED
  GET   /v3/resource-inventory/jobs/{id}/results → download JSON payload
  pandas + openpyxl                   → build formatted Excel workbook
  win32com Outlook                    → send email with attachment

Known API Limitations Handled
------------------------------
  1. State / Launch Date may be null  → replaced with "Not Available"
  2. Dimension keys must be internal  → configured in DIMENSION_KEYS below
  3. win32com requires active session → guards + timeout added
  4. Async API                        → polling loop with back-off
"""

import os
import sys
import time
import logging
import smtplib
from datetime import datetime, timezone
from pathlib import Path
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

import requests
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────
# 1.  CONFIGURATION  ← edit this block only
# ─────────────────────────────────────────────

API_KEY = os.environ.get("CLOUDABILITY_API_KEY", "YOUR_API_KEY_HERE")

BASE_URL = "https://app.cloudability.com"

# ── Dimension Keys ───────────────────────────
# These are the *internal* Cloudability dimension keys, NOT the display names.
# Open Cloudability → Settings → Tag Mapping to confirm your exact keys.
# Common defaults are shown; adjust if your org uses different mappings.
DIMENSION_KEYS = {
    "account_name":    "vendor_account_name",   # Account Name
    "region":          "region",                 # AWS Region
    "resource_id":     "resource_identifier",    # Resource ID
    "product_name":    "product_name",           # Product / SKU name
    "instance_name":   "extended_resource_name", # AWS 'Name' tag  ← verify this
    "environment":     "tag2",                   # Your Environment tag  ← verify this
    "state":           "resource_state",         # EC2 state (may return null)
    "launch_date":     "resource_creation_date", # Launch date   (may return null)
}

# The API filter: only EC2 resources
SERVICE_FILTER = {
    "field":    "product_name",
    "operator": "==",
    "value":    "Amazon Elastic Compute Cloud"
}

# ── Polling settings ─────────────────────────
POLL_INTERVAL_SECONDS = 10   # seconds between status checks
POLL_MAX_WAIT_SECONDS = 600  # give up after 10 minutes

# ── Output file ──────────────────────────────
OUTPUT_DIR  = Path(__file__).parent
REPORT_DATE = datetime.now(timezone.utc).strftime("%Y-%m-%d")
OUTPUT_FILE = OUTPUT_DIR / f"EC2_Resource_Inventory_{REPORT_DATE}.xlsx"

# ── Email settings ───────────────────────────
# Recipients list – the script sends a DRAFT in Outlook so you can
# review before sending.  Set AUTO_SEND = True to skip the draft step.
EMAIL_RECIPIENTS = [
    "team@yourcompany.com",        # ← replace
]
EMAIL_CC          = []             # optional CC list
EMAIL_SUBJECT     = f"EC2 Resource Inventory – {REPORT_DATE}"
EMAIL_BODY_HTML   = f"""
<html><body>
<p>Hi team,</p>
<p>Please find attached the EC2 Resource Inventory report for <b>{REPORT_DATE}</b>,
   pulled directly from Apptio Cloudability.</p>
<p>The report includes: Account Name, Region, Resource ID, Product Name,
   Instance Name, Environment, State, and Launch Date.</p>
<p><i>Note: "Not Available" in the State or Launch Date columns indicates that
   the Cloudability metadata poller has not yet captured that instance's data,
   or the cross-account IAM role does not grant metadata read permissions.</i></p>
<br><p>Regards,<br>Cloud FinOps Automation</p>
</body></html>
"""
AUTO_SEND = False   # False = create Outlook DRAFT; True = send immediately

# ─────────────────────────────────────────────
# 2.  LOGGING
# ─────────────────────────────────────────────

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.StreamHandler(sys.stdout),
        logging.FileHandler(OUTPUT_DIR / "cloudability_inventory.log"),
    ],
)
log = logging.getLogger(__name__)


# ─────────────────────────────────────────────
# 3.  CLOUDABILITY API CLIENT
# ─────────────────────────────────────────────

class CloudabilityClient:
    """Thin wrapper around the Cloudability Resource Inventory API."""

    def __init__(self, api_key: str):
        if not api_key or api_key == "YOUR_API_KEY_HERE":
            raise ValueError(
                "Set CLOUDABILITY_API_KEY environment variable or update API_KEY in the script."
            )
        self.session = requests.Session()
        # Cloudability uses HTTP Basic Auth: api_key as username, empty password
        self.session.auth = (api_key, "")
        self.session.headers.update({"Content-Type": "application/json"})

    # ── Step 1: Submit job ───────────────────

    def submit_job(self) -> str:
        """POST a new Resource Inventory export job and return the job ID."""
        payload = {
            "dimensions": list(DIMENSION_KEYS.values()),
            "filters": [SERVICE_FILTER],
            "date": {
                "type": "custom",
                # Resource Inventory is point-in-time; use today's date range
                "start": REPORT_DATE,
                "end":   REPORT_DATE,
            },
            "format": "json",
        }
        url = f"{BASE_URL}/v3/resource-inventory/jobs"
        log.info("Submitting Resource Inventory job to %s …", url)
        resp = self.session.post(url, json=payload, timeout=30)
        self._raise_for_status(resp, "submit_job")
        job_id = resp.json().get("id") or resp.json().get("job_id")
        log.info("Job submitted. Job ID: %s", job_id)
        return job_id

    # ── Step 2: Poll status ──────────────────

    def wait_for_completion(self, job_id: str) -> None:
        """Block until the job reaches FINISHED (or raise on timeout/failure)."""
        url      = f"{BASE_URL}/v3/resource-inventory/jobs/{job_id}"
        deadline = time.monotonic() + POLL_MAX_WAIT_SECONDS
        attempt  = 0

        while time.monotonic() < deadline:
            attempt += 1
            resp   = self.session.get(url, timeout=30)
            self._raise_for_status(resp, "poll_status")
            body   = resp.json()
            status = (body.get("status") or body.get("state") or "").upper()

            log.info("Poll #%d – status: %s", attempt, status)

            if status == "FINISHED":
                log.info("Job %s completed successfully.", job_id)
                return
            if status in ("FAILED", "ERROR", "CANCELLED"):
                raise RuntimeError(f"Cloudability job {job_id} ended with status: {status}")

            time.sleep(POLL_INTERVAL_SECONDS)

        raise TimeoutError(
            f"Job {job_id} did not finish within {POLL_MAX_WAIT_SECONDS} seconds."
        )

    # ── Step 3: Download results ─────────────

    def download_results(self, job_id: str) -> list[dict]:
        """GET the completed job's result payload and return as a list of dicts."""
        url  = f"{BASE_URL}/v3/resource-inventory/jobs/{job_id}/results"
        log.info("Downloading results from %s …", url)
        resp = self.session.get(url, timeout=120)
        self._raise_for_status(resp, "download_results")
        body = resp.json()

        # The API may wrap data in a 'results' or 'data' key
        if isinstance(body, list):
            records = body
        else:
            records = body.get("results") or body.get("data") or []

        log.info("Downloaded %d resource records.", len(records))
        return records

    @staticmethod
    def _raise_for_status(resp: requests.Response, step: str) -> None:
        if not resp.ok:
            raise requests.HTTPError(
                f"[{step}] HTTP {resp.status_code}: {resp.text[:500]}"
            )


# ─────────────────────────────────────────────
# 4.  DATA TRANSFORMATION
# ─────────────────────────────────────────────

# Maps internal dimension key → friendly column header in Excel
COLUMN_DISPLAY = {
    "vendor_account_name":    "Account Name",
    "region":                 "Region",
    "resource_identifier":    "Resource ID",
    "product_name":           "Product Name",
    "extended_resource_name": "Instance Name",
    "tag2":                   "Environment",
    "resource_state":         "State",
    "resource_creation_date": "Launch Date",
}


def records_to_dataframe(records: list[dict]) -> pd.DataFrame:
    """Flatten raw API records into a clean DataFrame with display headers."""
    if not records:
        log.warning("No records returned – DataFrame will be empty.")
        return pd.DataFrame(columns=list(COLUMN_DISPLAY.values()))

    rows = []
    for rec in records:
        row = {}
        for dim_key, display_name in COLUMN_DISPLAY.items():
            # API may nest values under 'dimensions' sub-dict or at top level
            value = (
                rec.get("dimensions", {}).get(dim_key)
                or rec.get(dim_key)
            )
            # ── Limitation 1: null handling for State & Launch Date ──────────
            if value is None or str(value).strip() in ("", "null", "None"):
                value = "Not Available"
            row[display_name] = value
        rows.append(row)

    df = pd.DataFrame(rows, columns=list(COLUMN_DISPLAY.values()))
    log.info("DataFrame shape: %s", df.shape)
    return df


# ─────────────────────────────────────────────
# 5.  EXCEL EXPORT  (openpyxl formatting)
# ─────────────────────────────────────────────

HEADER_FILL  = PatternFill("solid", start_color="1F4E79")   # dark navy
HEADER_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
DATA_FONT    = Font(name="Arial", size=10)
ALT_ROW_FILL = PatternFill("solid", start_color="DCE6F1")    # light blue
BORDER_SIDE  = Side(style="thin", color="B8CCE4")
CELL_BORDER  = Border(
    left=BORDER_SIDE, right=BORDER_SIDE,
    top=BORDER_SIDE,  bottom=BORDER_SIDE,
)

# Columns where "Not Available" should be highlighted amber
NULLABLE_COLUMNS = {"State", "Launch Date"}


def export_to_excel(df: pd.DataFrame, path: Path) -> None:
    """Write the DataFrame to a formatted Excel workbook."""
    df.to_excel(path, index=False, sheet_name="EC2 Inventory")
    wb = load_workbook(path)
    ws = wb.active

    # ── Header row styling ───────────────────
    for cell in ws[1]:
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = CELL_BORDER
    ws.row_dimensions[1].height = 30

    # ── Data rows ────────────────────────────
    col_names = df.columns.tolist()
    null_col_indices = {
        col_names.index(c) + 1  # 1-based
        for c in NULLABLE_COLUMNS if c in col_names
    }

    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        is_alt = (row_idx % 2 == 0)
        for cell in row:
            cell.font      = DATA_FONT
            cell.border    = CELL_BORDER
            cell.alignment = Alignment(vertical="center")
            if is_alt:
                cell.fill = ALT_ROW_FILL
            # Highlight "Not Available" cells amber
            if cell.column in null_col_indices and cell.value == "Not Available":
                cell.fill = PatternFill("solid", start_color="FFD966")

    # ── Auto-fit column widths ───────────────
    for col_cells in ws.columns:
        max_len = max(
            (len(str(c.value)) for c in col_cells if c.value is not None),
            default=10,
        )
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max_len + 4, 45)

    # ── Freeze header + add auto-filter ─────
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # ── Summary tab ──────────────────────────
    ws_sum = wb.create_sheet("Summary")
    ws_sum["A1"] = "Report Date"
    ws_sum["B1"] = REPORT_DATE
    ws_sum["A2"] = "Total EC2 Resources"
    ws_sum["B2"] = f"=COUNTA('EC2 Inventory'!A:A)-1"
    ws_sum["A3"] = "Records with State 'Not Available'"
    ws_sum["B3"] = f"=COUNTIF('EC2 Inventory'!G:G,\"Not Available\")"
    ws_sum["A4"] = "Records with Launch Date 'Not Available'"
    ws_sum["B4"] = f"=COUNTIF('EC2 Inventory'!H:H,\"Not Available\")"

    for r in range(1, 5):
        ws_sum.cell(r, 1).font = Font(name="Arial", bold=True, size=10)
        ws_sum.cell(r, 2).font = Font(name="Arial", size=10)
    ws_sum.column_dimensions["A"].width = 38
    ws_sum.column_dimensions["B"].width = 20

    wb.save(path)
    log.info("Excel report saved → %s", path)


# ─────────────────────────────────────────────
# 6.  EMAIL DISPATCH  (Outlook COM)
# ─────────────────────────────────────────────

def send_via_outlook(attachment_path: Path) -> None:
    """
    Create an Outlook mail item via win32com and either save as Draft
    or send immediately based on AUTO_SEND.

    Limitations guarded here:
      • Active session check  (headless environments will ImportError)
      • Hung Outlook guard    (COM dispatch timeout via threading)
    """
    try:
        import win32com.client
    except ImportError:
        log.error(
            "win32com not available.  This script must run on a Windows machine "
            "with pywin32 installed (`pip install pywin32`) and an active user session."
        )
        _fallback_smtp(attachment_path)
        return

    import threading

    result = {}
    error  = {}

    def _dispatch():
        try:
            outlook = win32com.client.Dispatch("Outlook.Application")
            mail    = outlook.CreateItem(0)   # 0 = olMailItem

            mail.Subject    = EMAIL_SUBJECT
            mail.HTMLBody   = EMAIL_BODY_HTML
            mail.To         = "; ".join(EMAIL_RECIPIENTS)
            if EMAIL_CC:
                mail.CC     = "; ".join(EMAIL_CC)

            mail.Attachments.Add(str(attachment_path.resolve()))

            if AUTO_SEND:
                mail.Send()
                result["status"] = "sent"
            else:
                mail.Save()   # saves to Drafts folder
                result["status"] = "draft"
        except Exception as exc:
            error["msg"] = str(exc)

    t = threading.Thread(target=_dispatch, daemon=True)
    t.start()
    t.join(timeout=60)   # wait up to 60 s; protects against hung OUTLOOK.EXE

    if t.is_alive():
        log.error(
            "Outlook COM call timed out after 60 seconds.  "
            "OUTLOOK.EXE may be in a hung/locked state – check for modal dialogs.  "
            "Falling back to SMTP."
        )
        _fallback_smtp(attachment_path)
    elif "msg" in error:
        log.error("Outlook COM error: %s.  Falling back to SMTP.", error["msg"])
        _fallback_smtp(attachment_path)
    else:
        action = "sent" if result.get("status") == "sent" else "saved to Drafts"
        log.info("Email %s successfully via Outlook.", action)


def _fallback_smtp(attachment_path: Path) -> None:
    """
    Fallback: write the report path to a text file so the operator
    knows where to find it.  Swap in SMTP credentials if you have a
    relay server available without Outlook.
    """
    notice = OUTPUT_DIR / "EMAIL_PENDING.txt"
    notice.write_text(
        f"Outlook COM dispatch failed.\n"
        f"Please manually attach and send:\n\n"
        f"  {attachment_path.resolve()}\n\n"
        f"Recipients : {', '.join(EMAIL_RECIPIENTS)}\n"
        f"Subject    : {EMAIL_SUBJECT}\n"
    )
    log.warning("Outlook unavailable. Instructions written to: %s", notice)


# ─────────────────────────────────────────────
# 7.  ORCHESTRATION
# ─────────────────────────────────────────────

def run() -> None:
    log.info("=" * 60)
    log.info("EC2 Resource Inventory Automation – %s", REPORT_DATE)
    log.info("=" * 60)

    # ── Step A: Validate config ──────────────
    if API_KEY == "YOUR_API_KEY_HERE":
        raise ValueError(
            "API key not set.  Export CLOUDABILITY_API_KEY=<your_key> "
            "or update API_KEY in the script."
        )

    client = CloudabilityClient(API_KEY)

    # ── Step B: Fetch data (async 3-step) ───
    job_id  = client.submit_job()
    client.wait_for_completion(job_id)
    records = client.download_results(job_id)

    # ── Step C: Transform → DataFrame ───────
    df = records_to_dataframe(records)
    log.info("Sample data:\n%s", df.head(3).to_string())

    # ── Step D: Export → Excel ───────────────
    export_to_excel(df, OUTPUT_FILE)

    # ── Step E: Email dispatch ───────────────
    send_via_outlook(OUTPUT_FILE)

    log.info("Pipeline complete.  Output: %s", OUTPUT_FILE)


if __name__ == "__main__":
    run()
